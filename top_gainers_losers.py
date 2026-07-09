"""
Top Gainers & Losers - Last 2 Months (Day by Day)
====================================================
Uses OpenAlgo's internal history service to fetch daily OHLC
for each stock in the watchlist and ranks them by % daily change.

Usage:
    python top_gainers_losers.py

Output:
    - top_gainers_losers.xlsx  (summary + per-day sheets)
    - top_gainers_losers.csv   (flat CSV with all daily data)
"""

import os
import sys
import time
import warnings
from datetime import date, timedelta, datetime
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Bootstrap OpenAlgo environment ──────────────────────────────────────────
# Load .env so services pick up DATABASE_URL, etc.
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

# ── Config ───────────────────────────────────────────────────────────────────
WATCHLIST_FILE = Path("/Users/bond7/Downloads/Future Stocks watchlist.txt")
API_KEY = os.getenv("APP_KEY", "")          # we'll read from .env / generate
TOP_N = 10                                   # top N gainers and losers per day
MONTHS_BACK = 2
EXCHANGE = "NSE"
INTERVAL = "D"                               # Daily candles
DELAY_BETWEEN_CALLS = 0.05                   # seconds between API calls (rate-limit friendly)
OUTPUT_EXCEL = Path(__file__).parent / "top_gainers_losers.xlsx"
OUTPUT_CSV = Path(__file__).parent / "top_gainers_losers.csv"

# ── Read watchlist ────────────────────────────────────────────────────────────
def load_watchlist() -> list[str]:
    raw = WATCHLIST_FILE.read_text().strip()
    symbols = []
    for token in raw.split(","):
        token = token.strip()
        if ":" in token:
            sym = token.split(":", 1)[1]
        else:
            sym = token
        sym = sym.strip()
        if sym:
            symbols.append(sym)
    print(f"[Watchlist] Loaded {len(symbols)} symbols")
    return symbols


# ── Date range ────────────────────────────────────────────────────────────────
def get_date_range() -> tuple[str, str]:
    end = date.today()
    # Go back ~2 months (≈62 trading days buffer)
    start = end - timedelta(days=62)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


# ── Fetch history via OpenAlgo internal service ───────────────────────────────
def fetch_history_for_symbol(symbol: str, start_date: str, end_date: str) -> list[dict] | None:
    """
    Call OpenAlgo's get_history() directly (no HTTP overhead).
    Returns list of OHLC dicts or None on error.
    """
    try:
        from services.history_service import get_history
        # We need the API key from DB — use the app key which is the login key
        # Alternatively fetch directly using broker service
        success, data, status = get_history(
            symbol=symbol,
            exchange=EXCHANGE,
            interval=INTERVAL,
            start_date=start_date,
            end_date=end_date,
            api_key=_get_api_key(),
            source="api",
        )
        if success and isinstance(data, dict) and data.get("status") == "success":
            return data.get("data", [])
        return None
    except Exception as e:
        print(f"  [WARN] {symbol}: {e}")
        return None


_cached_api_key = None

def _get_api_key() -> str:
    """
    Get the first active API key from the OpenAlgo database.
    Decrypts directly using Fernet (same as database/auth_db.py) — no Flask app needed.
    """
    global _cached_api_key
    if _cached_api_key:
        return _cached_api_key

    try:
        import base64
        from cryptography.fernet import Fernet
        from cryptography.hazmat.primitives import hashes
        from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
        from database.flow_db import db_session
        from sqlalchemy import text

        # Build Fernet key from API_KEY_PEPPER (same as auth_db.py)
        pepper = os.getenv("API_KEY_PEPPER", "")
        if not pepper:
            raise RuntimeError("API_KEY_PEPPER not set in .env")

        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=b"openalgo_static_salt",
            iterations=100000,
        )
        fernet_key = base64.urlsafe_b64encode(kdf.derive(pepper.encode()))
        fernet = Fernet(fernet_key)

        with db_session() as s:
            row = s.execute(text('SELECT api_key_encrypted FROM api_keys LIMIT 1')).fetchone()
            if row and row[0]:
                _cached_api_key = fernet.decrypt(row[0].encode()).decode()
                print(f"[Auth] API key decrypted successfully: {_cached_api_key[:8]}...")
                return _cached_api_key

    except Exception as e:
        print(f"[WARN] Direct decryption failed: {e}")

    raise RuntimeError(
        "Could not find or decrypt a valid API key in the database.\n"
        "Please log into OpenAlgo at http://127.0.0.1:5000 first, "
        "then re-run this script."
    )


# ── Build daily gains DataFrame ───────────────────────────────────────────────
def build_daily_df(symbols: list[str], start_date: str, end_date: str):
    import pandas as pd

    all_rows = []
    total = len(symbols)

    for i, sym in enumerate(symbols, 1):
        print(f"  [{i:>3}/{total}] Fetching {sym}...", end="", flush=True)
        candles = fetch_history_for_symbol(sym, start_date, end_date)
        if not candles:
            print(" ✗ no data")
            continue

        # Each candle: {"date": "YYYY-MM-DD", "open": x, "high": h, "low": l, "close": c, "volume": v}
        for c in candles:
            raw_date = c.get("date") or c.get("timestamp") or c.get("time")
            if not raw_date:
                continue
            # Normalise to YYYY-MM-DD string
            if isinstance(raw_date, (int, float)):
                raw_date = datetime.utcfromtimestamp(raw_date / 1000 if raw_date > 1e10 else raw_date).strftime("%Y-%m-%d")
            elif "T" in str(raw_date):
                raw_date = str(raw_date)[:10]
            else:
                raw_date = str(raw_date)[:10]

            open_p = float(c.get("open", 0) or 0)
            close_p = float(c.get("close", 0) or 0)
            high_p = float(c.get("high", 0) or 0)
            low_p = float(c.get("low", 0) or 0)
            volume = int(c.get("volume", 0) or 0)

            if open_p <= 0 or close_p <= 0:
                continue

            pct_change = ((close_p - open_p) / open_p) * 100

            all_rows.append({
                "date": raw_date,
                "symbol": sym,
                "open": open_p,
                "high": high_p,
                "low": low_p,
                "close": close_p,
                "volume": volume,
                "pct_change": round(pct_change, 2),
            })

        print(f" ✓ {len(candles)} candles")
        time.sleep(DELAY_BETWEEN_CALLS)

    if not all_rows:
        print("\n[ERROR] No data fetched. Make sure OpenAlgo is running and you are logged in.")
        sys.exit(1)

    df = pd.DataFrame(all_rows)
    df["date"] = pd.to_datetime(df["date"])
    df.sort_values(["date", "pct_change"], ascending=[True, False], inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


# ── Top N gainers / losers per day ────────────────────────────────────────────
def compute_daily_top(df, top_n: int = 10):
    import pandas as pd

    gainers_frames = []
    losers_frames = []

    for trade_date, group in df.groupby("date"):
        group_sorted = group.sort_values("pct_change", ascending=False)
        top_gain = group_sorted.head(top_n).copy()
        top_loss = group_sorted.tail(top_n).copy()
        top_gain["rank"] = range(1, len(top_gain) + 1)
        top_loss["rank"] = range(1, len(top_loss) + 1)
        top_gain["category"] = "Gainer"
        top_loss["category"] = "Loser"
        gainers_frames.append(top_gain)
        losers_frames.append(top_loss)

    gainers = pd.concat(gainers_frames, ignore_index=True)
    losers = pd.concat(losers_frames, ignore_index=True)
    return gainers, losers


# ── Excel export ──────────────────────────────────────────────────────────────
def export_excel(df_all: "pd.DataFrame", gainers: "pd.DataFrame", losers: "pd.DataFrame"):
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("[WARN] openpyxl not found. Install with: pip install openpyxl")
        print("       Skipping Excel export, CSV only.")
        return

    import pandas as pd
    from openpyxl.styles import (
        Alignment, Font, PatternFill, Border, Side
    )
    from openpyxl.utils import get_column_letter

    print("\n[Excel] Writing top_gainers_losers.xlsx ...")

    writer = pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl")

    # ── Sheet 1: Summary (Gainers) ──
    gain_pivot = gainers.pivot_table(
        index="date", columns="rank",
        values=["symbol", "pct_change"],
        aggfunc="first"
    )
    gain_pivot.columns = [f"Gainer_{c[1]}_{c[0]}" for c in gain_pivot.columns]
    gain_pivot.index = gain_pivot.index.strftime("%Y-%m-%d")
    gain_pivot.to_excel(writer, sheet_name="Top Gainers")

    # ── Sheet 2: Summary (Losers) ──
    loss_pivot = losers.pivot_table(
        index="date", columns="rank",
        values=["symbol", "pct_change"],
        aggfunc="first"
    )
    loss_pivot.columns = [f"Loser_{c[1]}_{c[0]}" for c in loss_pivot.columns]
    loss_pivot.index = loss_pivot.index.strftime("%Y-%m-%d")
    loss_pivot.to_excel(writer, sheet_name="Top Losers")

    # ── Sheet 3: Daily Detail (all data) ──
    detail = df_all.copy()
    detail["date"] = detail["date"].dt.strftime("%Y-%m-%d")
    detail.to_excel(writer, sheet_name="All Daily Data", index=False)

    # ── Sheet 4: Per-day gainers+losers side by side ──
    days = sorted(df_all["date"].unique())
    rows_combined = []
    for d in days:
        d_str = d.strftime("%Y-%m-%d")
        g_day = gainers[gainers["date"] == d].sort_values("rank")
        l_day = losers[losers["date"] == d].sort_values("rank")
        for rank in range(1, TOP_N + 1):
            grow = g_day[g_day["rank"] == rank]
            lrow = l_day[l_day["rank"] == rank]
            rows_combined.append({
                "date": d_str,
                "rank": rank,
                "gainer_symbol": grow["symbol"].values[0] if len(grow) else "",
                "gainer_%chg": grow["pct_change"].values[0] if len(grow) else "",
                "gainer_close": grow["close"].values[0] if len(grow) else "",
                "loser_symbol": lrow["symbol"].values[0] if len(lrow) else "",
                "loser_%chg": lrow["pct_change"].values[0] if len(lrow) else "",
                "loser_close": lrow["close"].values[0] if len(lrow) else "",
            })

    combined_df = pd.DataFrame(rows_combined)
    combined_df.to_excel(writer, sheet_name="Daily Summary", index=False)

    writer.close()

    # ── Apply colour formatting ──
    wb = writer.book if hasattr(writer, "book") else __import__("openpyxl").load_workbook(OUTPUT_EXCEL)
    # Re-open and style
    wb = __import__("openpyxl").load_workbook(OUTPUT_EXCEL)

    GREEN_HEADER = PatternFill("solid", fgColor="1A7A4A")
    RED_HEADER   = PatternFill("solid", fgColor="B22222")
    BLUE_HEADER  = PatternFill("solid", fgColor="1E3A5F")
    LIGHT_GREEN  = PatternFill("solid", fgColor="E8F5E9")
    LIGHT_RED    = PatternFill("solid", fgColor="FFEBEE")
    WHITE_FONT   = Font(color="FFFFFF", bold=True)
    BOLD         = Font(bold=True)

    def style_header_row(ws, fill, font=WHITE_FONT):
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    # Style "Top Gainers"
    ws_g = wb["Top Gainers"]
    style_header_row(ws_g, GREEN_HEADER)
    auto_width(ws_g)

    # Style "Top Losers"
    ws_l = wb["Top Losers"]
    style_header_row(ws_l, RED_HEADER)
    auto_width(ws_l)

    # Style "All Daily Data" — colour rows by pct_change
    ws_a = wb["All Daily Data"]
    style_header_row(ws_a, BLUE_HEADER)
    pct_col_idx = None
    for j, cell in enumerate(ws_a[1], 1):
        if str(cell.value).lower() in ("pct_change", "% change"):
            pct_col_idx = j
            break
    if pct_col_idx:
        for row in ws_a.iter_rows(min_row=2):
            pct_val = row[pct_col_idx - 1].value
            if pct_val is not None:
                fill = LIGHT_GREEN if float(pct_val) >= 0 else LIGHT_RED
                for cell in row:
                    cell.fill = fill
    auto_width(ws_a)

    # Style "Daily Summary"
    ws_d = wb["Daily Summary"]
    style_header_row(ws_d, BLUE_HEADER)
    # Colour gainer cols green, loser cols red
    for row in ws_d.iter_rows(min_row=2):
        for cell in row:
            col_name = ws_d.cell(1, cell.column).value or ""
            if "gainer" in str(col_name).lower():
                cell.fill = LIGHT_GREEN
            elif "loser" in str(col_name).lower():
                cell.fill = LIGHT_RED
    auto_width(ws_d)

    wb.save(OUTPUT_EXCEL)
    print(f"[Excel] Saved → {OUTPUT_EXCEL}")


# ── CSV export ────────────────────────────────────────────────────────────────
def export_csv(gainers: "pd.DataFrame", losers: "pd.DataFrame"):
    import pandas as pd

    combined = pd.concat([gainers, losers], ignore_index=True)
    combined["date"] = combined["date"].dt.strftime("%Y-%m-%d")
    combined.sort_values(["date", "category", "rank"], inplace=True)
    combined.to_csv(OUTPUT_CSV, index=False)
    print(f"[CSV]   Saved → {OUTPUT_CSV}")


# ── Print summary to console ──────────────────────────────────────────────────
def print_summary(gainers: "pd.DataFrame", losers: "pd.DataFrame"):
    import pandas as pd

    days = sorted(gainers["date"].unique())
    print(f"\n{'='*70}")
    print(f"  TOP {TOP_N} GAINERS & LOSERS — {days[0].date()} to {days[-1].date()}")
    print(f"{'='*70}")

    for d in days[-5:]:   # show last 5 trading days
        d_str = d.strftime("%Y-%m-%d")
        g = gainers[gainers["date"] == d].sort_values("rank")
        l = losers[losers["date"] == d].sort_values("rank")
        print(f"\n  📅 {d_str}")
        print(f"  {'GAINERS':35s}  {'LOSERS':35s}")
        print(f"  {'-'*35}  {'-'*35}")
        for rank in range(1, TOP_N + 1):
            gr = g[g["rank"] == rank]
            lr = l[l["rank"] == rank]
            gsym = f"{gr['symbol'].values[0]} ({gr['pct_change'].values[0]:+.2f}%)" if len(gr) else ""
            lsym = f"{lr['symbol'].values[0]} ({lr['pct_change'].values[0]:+.2f}%)" if len(lr) else ""
            print(f"  {rank:>2}. {'🟢 ' + gsym:35s}  {'🔴 ' + lsym}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("  OpenAlgo — Top Gainers & Losers (Last 2 Months, Day by Day)")
    print("=" * 70)

    # 1. Load watchlist
    symbols = load_watchlist()

    # 2. Date range
    start_date, end_date = get_date_range()
    print(f"[Dates] {start_date}  →  {end_date}")

    # 3. Fetch all data
    print(f"\n[Fetch] Downloading daily OHLC for {len(symbols)} symbols...\n")
    df_all = build_daily_df(symbols, start_date, end_date)

    print(f"\n[Data] Total records: {len(df_all)} across {df_all['date'].nunique()} trading days")

    # 4. Compute daily top gainers / losers
    gainers, losers = compute_daily_top(df_all, top_n=TOP_N)

    # 5. Print summary
    print_summary(gainers, losers)

    # 6. Export CSV (always works)
    export_csv(gainers, losers)

    # 7. Export Excel (needs openpyxl)
    export_excel(df_all, gainers, losers)

    print("\n✅ Done!")
    print(f"   Excel → {OUTPUT_EXCEL}")
    print(f"   CSV   → {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
